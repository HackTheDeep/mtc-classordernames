import sys
import requests
import re
import json
from openpyxl import load_workbook


fileName = "dirtydataset.xlsx"


def main():
    if(sys.argv[1] == "classify"):
        classify()
    # add additional tasks here as command line arguments


def classify():
    startRow = sys.argv[2]
    endRow = sys.argv[3]
    print("Adding classes and orders to spreadsheet...")
    wb = load_workbook(fileName)
    ws = wb.active
    currentRow = startRow
    for row in ws.iter_rows(min_col=13, min_row=startRow, max_col=13, max_row=endRow):
        for cell in row:
            classOrderTuple = getClassAndOrder(cell.value)
            ws.cell(row=currentRow, column=9).value = classOrderTuple[0]
            ws.cell(row=currentRow, column=11).value = classOrderTuple[1]
        currentRow += 1
    wb.save(fileName)


def getClassAndOrder(family):
    specimenClass = ""
    specimenOrder = ""
    # Method 1: Scraping animaldiversity.org
    # probably not a good idea for thousands of entries
    #response = requests.get("http://animaldiversity.org/accounts/" + family + "/classification/").text
    #specimenClass = re.search('(?<=Class</span><a name=")[^"]+', response).group(0)
    #specimenOrder = re.search('(?<=Order</span><a name=")[^"]+', response).group(0)
    #return (specimenClass, specimenOrder)

    # Method 2: Using the WoRMS API
    response = requests.get("http://www.marinespecies.org/rest/AphiaRecordsByMatchNames?scientificnames%5B%5D=" + family + "&marine_only=true").json()
    specimenClass = response[0][0]['class']
    specimenOrder = response[0][0]['order']
    return (specimenClass, specimenOrder)

    # Method 3: Using the wikipedia API
    #wikiAPIRoot = "https://en.wikipedia.org/w/api.php?action=query&prop=revisions&rvprop=content&format=json&titles="
    #return(specimenClass, specimenOrder)


main()
